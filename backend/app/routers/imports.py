from fastapi import APIRouter, Depends, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.vehicle import Vehicle
from app.models.driver import Driver
from app.models.assignment import Assignment
import io

router = APIRouter(prefix="/api/import", tags=["import"])

SHEET_CATEGORY_MAP = {
    "Marketing & PR Fleet": "marketing_pr",
    "Demo Fleet": "demo",
    "De-Fleet": "de_fleet",
    "NAVS": "navs",
    "PTO Fleet": "pto",
    "IECP Vehicles": "iecp",
    "IECP Fleet": "iecp",
    "Internal Fleet": "internal",
}


def safe(val):
    if val is None:
        return ""
    return str(val).strip()


@router.post("/fleet-tracker")
async def import_fleet_tracker(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import openpyxl

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)

    stats = {"vehicles_imported": 0, "drivers_imported": 0, "sheets_processed": []}

    for sheet_name in wb.sheetnames:
        category = SHEET_CATEGORY_MAP.get(sheet_name)
        if not category:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [safe(h).lower().replace(" ", "_").replace("/", "_") for h in rows[0]]
        stats["sheets_processed"].append(sheet_name)

        for row in rows[1:]:
            data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            vin = safe(data.get("vin", data.get("vin_#", data.get("vin_number", ""))))
            if not vin or len(vin) < 5:
                continue

            existing = db.query(Vehicle).filter(Vehicle.vin == vin).first()
            if existing:
                existing.category = category
                existing.status = "returned" if category == "de_fleet" else "active"
            else:
                v = Vehicle(
                    vin=vin,
                    year=safe(data.get("year", data.get("model_year", ""))),
                    model=safe(data.get("model", "Grenadier")),
                    body_style=safe(data.get("body", data.get("body_style", ""))),
                    trim=safe(data.get("trim", "")),
                    color=safe(data.get("color", data.get("int_color", data.get("interior_color", "")))),
                    ext_color=safe(data.get("ext_color", data.get("exterior_color", ""))),
                    category=category,
                    status="returned" if category == "de_fleet" else "active",
                    location_city=safe(data.get("city", data.get("location", ""))),
                    location_state=safe(data.get("state", "")),
                    dealer_name=safe(data.get("dealer", data.get("dealership", data.get("dealer_name", "")))),
                    plate_number=safe(data.get("plate", data.get("plate_#", data.get("plate_number", "")))),
                    notes=safe(data.get("notes", data.get("note", ""))),
                )
                try:
                    v.mileage = int(float(safe(data.get("mileage", "0")) or "0"))
                except (ValueError, TypeError):
                    v.mileage = None
                db.add(v)
                stats["vehicles_imported"] += 1

            driver_name = safe(data.get("driver", data.get("primary_driver", data.get("name", ""))))
            if driver_name and len(driver_name) > 2:
                existing_driver = db.query(Driver).filter(Driver.name == driver_name).first()
                if not existing_driver:
                    d = Driver(
                        name=driver_name,
                        email=safe(data.get("email", "")),
                        department=safe(data.get("department", "")),
                    )
                    db.add(d)
                    stats["drivers_imported"] += 1

    db.commit()
    return stats
